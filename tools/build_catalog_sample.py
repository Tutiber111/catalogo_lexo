from __future__ import annotations

import json
import re
from pathlib import Path

import fitz
from openpyxl import load_workbook
from PIL import Image


ROOT = Path(__file__).resolve().parents[1]
PDF_BASE = Path(r"C:\Users\Lenovo\Dropbox\ACCESO A CLIENTES")
DOWNLOADS_BASE = Path(r"C:\Users\Lenovo\Downloads")
EXCEL_PATH = Path(r"C:\Users\Lenovo\Desktop\Listas de precios\Lista Lexo - Abril 2026.xlsx")
WEB_DIR = ROOT / "web"
PAGE_DIR = WEB_DIR / "assets" / "pages"
DATA_DIR = WEB_DIR / "data"

SELECTED_PAGES = [
    *range(1, 17),
    20,
    30,
    40,
    50,
    75,
    100,
    125,
    150,
    200,
]
LEXO_PAGES = range(1, 26)
ESTIA_PAGES = range(26, 51)
MAGEFESA_PAGES = range(51, 68)

SKU_ALIASES = {
    "40004": "400004",
    "40006": "400006",
    "2550": "2551",
}

PRICE_RE = re.compile(r"^\$[\d.]+(?:,\d+)?$")
SKU_RE = re.compile(r"^\d{3,6}$")
SIZE_RE = re.compile(r"(?P<value>\d+(?:[,.]\d+)?)\s*(?P<unit>ml|l|litro|litros)\b", re.IGNORECASE)

PRICE_POSITION_OVERRIDES = {
    # Multi-variant bottle/thermo pages: align prices to the capacity rows.
    (4, "500ML", "$13.057"): {"x": 0.89, "y": 0.690},
    (4, "750ML", "$14.820"): {"x": 0.89, "y": 0.718},
    (13, "500ML", "$11.187"): {"x": 0.88, "y": 0.630},
    (13, "1 Litro", "$15.743"): {"x": 0.88, "y": 0.658},
    # Cafeteras and accessories: place prices below the capacity/cup-count labels.
    (15, "Cafetera French Press 350ml", "$8.843"): {"x": 0.184, "y": 0.760},
    (15, "Cafetera French Press 600ml", "$11.990"): {"x": 0.438, "y": 0.760},
    (15, "Cafetera French Press 1000ml", "$14.712"): {"x": 0.691, "y": 0.760},
    (16, "350ML", "$8.843"): {"x": 0.478, "y": 0.565},
    (16, "1 Litro", "$14.712"): {"x": 0.478, "y": 0.615},
    (17, "Cafetera Acero French Press 350ml", "$13.739"): {"x": 0.258, "y": 0.835},
    (17, "Cafetera Acero French Press 1000ml", "$20.189"): {"x": 0.562, "y": 0.835},
    (18, "Cafetera Aluminio 3 Tazas Marca Lexo", "$11.935"): {"x": 0.091, "y": 0.720},
    (18, "Cafetera Aluminio 6 Tazas Marca Lexo", "$15.017"): {"x": 0.328, "y": 0.720},
    (18, "Cafetera Aluminio 9 Tazas Marca Lexo", "$20.429"): {"x": 0.565, "y": 0.720},
    (18, "Cafetera Aluminio 12 Tazas Marca Lexo", "$24.079"): {"x": 0.810, "y": 0.720},
    (19, "Cafetera Acero 4 Tazas Marca Lexo", "$19.679"): {"x": 0.189, "y": 0.695},
    (19, "Cafetera Acero 6 Tazas Marca Lexo", "$22.825"): {"x": 0.441, "y": 0.695},
    (19, "Cafetera Acero 9 Tazas Marca Lexo", "$28.578"): {"x": 0.694, "y": 0.695},
    (22, "Tetera 800Ml", "$20.866"): {"x": 0.240, "y": 0.635},
    (22, "Tetera 1 Litro", "$18.245"): {"x": 0.743, "y": 0.635},
    (23, "Jarra Medidora 500Ml", "$12.094"): {"x": 0.500, "y": 0.695},
    (24, "Pack X 2 Vasos Dobles 80Ml", "$12.457"): {"x": 0.330, "y": 0.680},
    (24, "Pack X 2 Vasos Dobles 250Ml", "$16.247"): {"x": 0.670, "y": 0.680},
    (25, "Frasco Con Tapa De Bamboo 450Ml", "$7.578"): {"x": 0.068, "y": 0.688},
    (25, "Frasco Con Tapa De Bamboo 800Ml", "$8.279"): {"x": 0.306, "y": 0.688},
    (25, "Frasco Con Tapa De Bamboo 1 litro", "$10.641"): {"x": 0.546, "y": 0.688},
    (25, "Frasco Con Tapa De Bamboo 1.3 litros", "$11.574"): {"x": 0.808, "y": 0.688},
}


def find_pdf() -> Path:
    matches = list(DOWNLOADS_BASE.glob("Copy of Cat*logo Lexo.pdf"))
    if matches:
        return matches[0]
    matches = list(PDF_BASE.glob("Cat*logos de productos/Cat*logo Completo.pdf"))
    if not matches:
        raise FileNotFoundError("Could not find catalog PDF under Dropbox access folder.")
    return matches[0]


def find_priced_reference_pdf() -> Path | None:
    matches = list(PDF_BASE.glob("Cat*logos de productos/Cat*logo Completo.pdf"))
    return matches[0] if matches else None


def clean_text(value: str) -> str:
    return " ".join(value.replace("\n", " ").split()).strip()


def normalize_sku(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return clean_text(str(value))


def format_price(value) -> str:
    if not isinstance(value, (int, float)):
        return ""
    rounded = int(round(value))
    return "$" + f"{rounded:,}".replace(",", ".")


def normalize_size_label(value: str) -> str:
    text = clean_text(value).replace(",", ".")
    match = SIZE_RE.search(text)
    if not match:
        return ""
    number = match.group("value").replace(",", ".")
    unit = match.group("unit").lower()
    try:
        numeric = float(number)
    except ValueError:
        numeric = 0

    if unit in {"l", "litro", "litros"}:
        if numeric == 1:
            return "1 Litro"
        return f"{numeric:g} Litros"
    if numeric == 1000:
        return "1 Litro"
    return f"{numeric:g}ML"


def product_size_label(description: str) -> str:
    return normalize_size_label(description)


def size_value_ml(label: str) -> float | None:
    match = SIZE_RE.search(clean_text(label).replace(",", "."))
    if not match:
        return None
    try:
        value = float(match.group("value").replace(",", "."))
    except ValueError:
        return None
    unit = match.group("unit").lower()
    if unit in {"l", "litro", "litros"}:
        return value * 1000
    return value


def load_price_list() -> dict[str, dict]:
    if not EXCEL_PATH.exists():
        return {}

    workbook = load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    products = {}

    for row_number, row in enumerate(sheet.iter_rows(min_col=1, max_col=6, values_only=True), start=1):
        code = normalize_sku(row[1])
        description = clean_text(str(row[2])) if row[2] else ""
        if not code or not description or code.lower() in {"código", "codigo"}:
            continue

        price_value = row[5]
        products[code] = {
            "sku": code,
            "description": description,
            "ean": normalize_sku(row[3]),
            "unitsPerCase": row[4] if isinstance(row[4], (int, float)) else None,
            "priceValue": price_value if isinstance(price_value, (int, float)) else None,
            "price": format_price(price_value),
            "sourceRow": row_number,
        }

    return products


def word_text(word: tuple) -> str:
    return str(word[4]).strip()


def find_category(blocks: list[tuple]) -> str:
    candidates = []
    for block in blocks:
        x0, y0, x1, y1, text = block[:5]
        label = clean_text(text)
        if not label or y0 > 210:
            continue
        if "$" in label or SKU_RE.fullmatch(label):
            continue
        score = (x1 - x0) * (y1 - y0)
        candidates.append((y0, -score, label))
    if not candidates:
        return "Catalog"
    return sorted(candidates)[0][2]


def words_on_line(words: list[tuple], target: tuple, tolerance: float = 8) -> list[tuple]:
    ty0, ty1 = target[1], target[3]
    middle = (ty0 + ty1) / 2
    line = [w for w in words if abs(((w[1] + w[3]) / 2) - middle) <= tolerance]
    return sorted(line, key=lambda w: w[0])


def nearest_label(blocks: list[tuple], price_word: tuple, category: str) -> str:
    px = (price_word[0] + price_word[2]) / 2
    py = (price_word[1] + price_word[3]) / 2
    candidates = []
    for block in blocks:
        x0, y0, x1, y1, text = block[:5]
        label = clean_text(text)
        if not label or label == category or "$" in label:
            continue
        if not any(ch.isalpha() for ch in label):
            continue
        if y0 > py + 40:
            continue
        bx = (x0 + x1) / 2
        by = (y0 + y1) / 2
        distance = abs(px - bx) * 0.5 + abs(py - by)
        candidates.append((distance, y0, label))
    if not candidates:
        return category
    return sorted(candidates)[0][2]


def parse_product_name(words: list[tuple], blocks: list[tuple], price_word: tuple, category: str) -> str:
    line = words_on_line(words, price_word)
    price_index = next((i for i, w in enumerate(line) if w == price_word), -1)
    after = [word_text(w) for w in line[price_index + 1 :] if not SKU_RE.fullmatch(word_text(w))]
    before = [word_text(w) for w in line[:price_index] if not SKU_RE.fullmatch(word_text(w))]
    inline = clean_text(" ".join(after or before))
    if inline and any(ch.isalpha() for ch in inline):
        return inline
    return nearest_label(blocks, price_word, category)


def nearest_skus(words: list[tuple], price_word: tuple) -> list[str]:
    px = (price_word[0] + price_word[2]) / 2
    py = (price_word[1] + price_word[3]) / 2
    candidates = []
    for word in words:
        text = word_text(word)
        if not SKU_RE.fullmatch(text):
            continue
        wx = (word[0] + word[2]) / 2
        wy = (word[1] + word[3]) / 2
        if abs(wy - py) > 180:
            continue
        distance = abs(px - wx) * 0.4 + abs(py - wy)
        candidates.append((distance, text))
    deduped = []
    for _, sku in sorted(candidates)[:4]:
        if sku not in deduped:
            deduped.append(sku)
    return deduped


def product_hotspot(page_rect: fitz.Rect, price_word: tuple, product_count: int) -> dict[str, float]:
    width = page_rect.width
    height = page_rect.height
    if product_count == 1:
        return {"x": 0.08, "y": 0.12, "w": 0.84, "h": 0.72}

    cx = ((price_word[0] + price_word[2]) / 2) / width
    cy = ((price_word[1] + price_word[3]) / 2) / height
    x = max(0.02, cx - 0.16)
    y = max(0.10, cy - 0.30)
    return {"x": x, "y": y, "w": min(0.32, 0.98 - x), "h": min(0.36, 0.92 - y)}


def sku_from_word(word: tuple, price_list: dict[str, dict]) -> str:
    text = normalize_sku(word_text(word)).strip(".,;:()[]")
    text = SKU_ALIASES.get(text, text)
    if text in price_list:
        return text
    return ""


def sku_hotspot(page_rect: fitz.Rect, sku_word: tuple) -> dict[str, float]:
    width = page_rect.width
    height = page_rect.height
    pad_x = 8
    pad_y = 5
    x = max(0.01, (sku_word[0] - pad_x) / width)
    y = max(0.01, (sku_word[1] - pad_y) / height)
    w = min(0.22, ((sku_word[2] - sku_word[0]) + pad_x * 2) / width)
    h = min(0.06, ((sku_word[3] - sku_word[1]) + pad_y * 2) / height)
    return {"x": x, "y": y, "w": min(w, 0.99 - x), "h": min(h, 0.99 - y)}


def sku_price_position(page_rect: fitz.Rect, sku_word: tuple) -> dict[str, float]:
    width = page_rect.width
    height = page_rect.height
    return {
        "x": min(0.94, max(0.06, ((sku_word[0] + sku_word[2]) / 2) / width)),
        "y": min(0.94, max(0.08, (sku_word[3] + 7) / height)),
    }


def extract_products_from_skus(page: fitz.Page, page_number: int, price_list: dict[str, dict]) -> list[dict]:
    words = page.get_text("words")
    blocks = page.get_text("blocks")
    category = find_category(blocks)
    products = []
    seen = set()

    for word in words:
        sku = sku_from_word(word, price_list)
        if not sku or sku in seen:
            continue
        seen.add(sku)
        data = price_list[sku]
        size_label = product_size_label(data["description"])
        products.append(
            {
                "id": f"p{page_number:03d}-{len(products) + 1}",
                "page": page_number,
                "sku": sku,
                "skus": [sku],
                "name": data["description"],
                "category": category,
                "price": data["price"],
                "pdfPrice": "",
                "priceSource": "excel",
                "ean": data["ean"],
                "unitsPerCase": data["unitsPerCase"],
                "sizeLabel": size_label,
                "hotspot": sku_hotspot(page.rect, word),
                "pricePosition": sku_price_position(page.rect, word),
            }
        )

    return products


def extract_products(page: fitz.Page, page_number: int, price_list: dict[str, dict]) -> list[dict]:
    blocks = page.get_text("blocks")
    words = page.get_text("words")
    category = find_category(blocks)
    price_words = [w for w in words if PRICE_RE.fullmatch(word_text(w))]
    if not price_words:
        return extract_products_from_skus(page, page_number, price_list)

    products = []
    for index, price_word in enumerate(price_words):
        price = word_text(price_word)
        name = parse_product_name(words, blocks, price_word, category)
        skus = nearest_skus(words, price_word)
        sku = skus[0] if skus else f"P{page_number:03d}-{index + 1}"
        price_match = next((price_list[item] for item in skus if item in price_list), None)
        size_label = product_size_label(price_match["description"] if price_match else name)
        products.append(
            {
                "id": f"p{page_number:03d}-{index + 1}",
                "page": page_number,
                "sku": sku,
                "skus": skus,
                "name": price_match["description"] if price_match else name,
                "category": category,
                "price": price_match["price"] if price_match and price_match["price"] else price,
                "pdfPrice": price,
                "priceSource": "excel" if price_match else "pdf",
                "ean": price_match["ean"] if price_match else "",
                "unitsPerCase": price_match["unitsPerCase"] if price_match else None,
                "sizeLabel": size_label,
                "hotspot": product_hotspot(page.rect, price_word, len(price_words)),
                "pricePosition": {
                    "x": min(0.94, max(0.06, ((price_word[0] + price_word[2]) / 2) / page.rect.width)),
                    "y": min(0.94, max(0.08, price_word[1] / page.rect.height)),
                },
            }
        )
    return products


def rect_center_position(page_rect: fitz.Rect, block: tuple, offset_y: float = 26) -> dict[str, float]:
    x0, y0, x1, y1 = block[:4]
    return {
        "x": min(0.94, max(0.06, ((x0 + x1) / 2) / page_rect.width)),
        "y": min(0.94, max(0.08, (y1 + offset_y) / page_rect.height)),
    }


def price_box(position: dict[str, float]) -> dict[str, float]:
    width = 0.14
    height = 0.035
    return {
        "x0": position["x"] - width / 2,
        "y0": position["y"],
        "x1": position["x"] + width / 2,
        "y1": position["y"] + height,
    }


def block_box(page_rect: fitz.Rect, block: tuple) -> dict[str, float]:
    x0, y0, x1, y1 = block[:4]
    return {"x0": x0 / page_rect.width, "y0": y0 / page_rect.height, "x1": x1 / page_rect.width, "y1": y1 / page_rect.height}


def boxes_overlap(a: dict[str, float], b: dict[str, float], margin: float = 0.008) -> bool:
    return not (
        a["x1"] < b["x0"] - margin
        or a["x0"] > b["x1"] + margin
        or a["y1"] < b["y0"] - margin
        or a["y0"] > b["y1"] + margin
    )


def collides_with_text(page: fitz.Page, position: dict[str, float], ignored_label: str = "") -> bool:
    box = price_box(position)
    ignored_size = normalize_size_label(ignored_label)
    for block in page.get_text("blocks"):
        text = clean_text(block[4])
        if not text:
            continue
        if ignored_size and normalize_size_label(text) == ignored_size:
            continue
        if boxes_overlap(box, block_box(page.rect, block)):
            return True
    return False


def avoid_text_collision(page: fitz.Page, base: dict[str, float], label: str = "") -> dict[str, float]:
    offsets = [0, 0.035, -0.045, 0.07, -0.08, 0.105, -0.115]
    for offset in offsets:
        candidate = {"x": base["x"], "y": min(0.94, max(0.08, base["y"] + offset))}
        if not collides_with_text(page, candidate, label):
            return candidate
    return base


def find_size_label_position(page: fitz.Page, label: str) -> dict[str, float] | None:
    if not label:
        return None
    target = normalize_size_label(label)
    candidates = []
    for block in page.get_text("blocks"):
        text = clean_text(block[4])
        if normalize_size_label(text) != target:
            continue
        has_sku = any(SKU_RE.fullmatch(part.strip(".,;:()[]")) for part in text.split())
        score = len(text) + (100 if has_sku else 0)
        candidates.append((score, block))
    if not candidates:
        return None
    return rect_center_position(page.rect, sorted(candidates, key=lambda item: item[0])[0][1])


def average_position(products: list[dict]) -> dict[str, float]:
    positions = [product.get("pricePosition") for product in products if product.get("pricePosition")]
    if not positions:
        return {"x": 0.5, "y": 0.5}
    return {
        "x": sum(pos["x"] for pos in positions) / len(positions),
        "y": sum(pos["y"] for pos in positions) / len(positions),
    }


def build_price_groups(page: fitz.Page, products: list[dict]) -> list[dict]:
    grouped: dict[tuple, list[dict]] = {}
    for product in products:
        key = (product.get("sizeLabel") or product["sku"], product["price"])
        grouped.setdefault(key, []).append(product)

    price_groups = []
    for index, ((size_label, price), items) in enumerate(grouped.items(), start=1):
        display_label = size_label if len(items) > 1 else items[0]["name"]
        position = find_size_label_position(page, size_label) if len(items) > 1 else None
        if position is None:
            position = average_position(items)
        position = avoid_text_collision(page, position, size_label)
        position = (
            PRICE_POSITION_OVERRIDES.get((items[0]["page"], size_label, price))
            or PRICE_POSITION_OVERRIDES.get((items[0]["page"], display_label, price))
            or PRICE_POSITION_OVERRIDES.get((items[0]["page"], items[0]["sku"], price))
            or position
        )
        price_groups.append(
            {
                "id": f"pg{items[0]['page']:03d}-{index}",
                "page": items[0]["page"],
                "label": display_label,
                "price": price,
                "productIds": [item["id"] for item in items],
                "position": position,
                "positionSource": "auto",
            }
        )
    return price_groups


def extract_reference_price_positions(reference_doc: fitz.Document | None, page_number: int) -> list[dict]:
    if reference_doc is None or page_number > reference_doc.page_count:
        return []
    page = reference_doc[page_number - 1]
    positions = []
    for word in page.get_text("words"):
        price = word_text(word)
        if not PRICE_RE.fullmatch(price):
            continue
        positions.append(
            {
                "price": price,
                "x": ((word[0] + word[2]) / 2) / page.rect.width,
                "y": word[1] / page.rect.height,
            }
        )
    return positions


def position_distance(a: dict, b: dict) -> float:
    return abs(a["x"] - b["x"]) + abs(a["y"] - b["y"])


def sort_reference_positions(positions: list[dict]) -> list[dict]:
    if not positions:
        return []
    x_range = max(pos["x"] for pos in positions) - min(pos["x"] for pos in positions)
    y_range = max(pos["y"] for pos in positions) - min(pos["y"] for pos in positions)
    if y_range > 0.015 and y_range >= x_range * 0.35:
        return sorted(positions, key=lambda pos: (pos["y"], pos["x"]))
    return sorted(positions, key=lambda pos: (pos["x"], pos["y"]))


def sort_groups_for_reference(groups: list[dict]) -> list[dict]:
    size_values = [size_value_ml(group["label"]) for group in groups]
    if groups and all(value is not None for value in size_values):
        return [group for _, group in sorted(zip(size_values, groups), key=lambda item: item[0])]
    return sorted(groups, key=lambda group: (group["position"]["y"], group["position"]["x"]))


def apply_reference_price_positions(price_groups: list[dict], reference_positions: list[dict]) -> None:
    if not price_groups or not reference_positions:
        return

    unused = reference_positions[:]
    for group in price_groups:
        exact = [pos for pos in unused if pos["price"] == group["price"]]
        if not exact:
            continue
        chosen = min(exact, key=lambda pos: position_distance(pos, group["position"]))
        group["position"] = {"x": chosen["x"], "y": chosen["y"]}
        group["positionSource"] = "priced-pdf"
        unused.remove(chosen)

    remaining_groups = [group for group in price_groups if group.get("positionSource") != "priced-pdf"]
    if not remaining_groups or not unused:
        return

    sorted_groups = sort_groups_for_reference(remaining_groups)
    sorted_refs = sort_reference_positions(unused)
    for group, ref in zip(sorted_groups, sorted_refs):
        group["position"] = {"x": ref["x"], "y": ref["y"]}
        group["positionSource"] = "priced-pdf-order"


def render_page(page: fitz.Page, destination: Path) -> dict[str, int | str]:
    pix = page.get_pixmap(matrix=fitz.Matrix(1.7, 1.7), alpha=False)
    image = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
    destination.parent.mkdir(parents=True, exist_ok=True)
    image.save(destination, "JPEG", quality=86, optimize=True)
    return {"src": str(destination.relative_to(WEB_DIR)).replace("\\", "/"), "width": pix.width, "height": pix.height}


def main() -> None:
    pdf = find_pdf()
    reference_pdf = find_priced_reference_pdf()
    price_list = load_price_list()
    doc = fitz.open(pdf)
    reference_doc = fitz.open(reference_pdf) if reference_pdf else None
    PAGE_DIR.mkdir(parents=True, exist_ok=True)
    DATA_DIR.mkdir(parents=True, exist_ok=True)

    pages = []
    products = []
    page_entries = []
    for page_number in LEXO_PAGES:
        if page_number <= doc.page_count:
            page_entries.append({"page": page_number, "doc": doc, "section": "Lexo", "showPriceOverlays": True})
    if reference_doc:
        for page_number in ESTIA_PAGES:
            if page_number <= reference_doc.page_count:
                page_entries.append({"page": page_number, "doc": reference_doc, "section": "Estia", "showPriceOverlays": False})
        for page_number in MAGEFESA_PAGES:
            if page_number <= reference_doc.page_count:
                page_entries.append({"page": page_number, "doc": reference_doc, "section": "Magefesa", "showPriceOverlays": False})

    for entry in page_entries:
        page_number = entry["page"]
        page_doc = entry["doc"]
        if page_number > page_doc.page_count:
            continue
        page = page_doc[page_number - 1]
        image_name = f"page-{page_number:03d}.jpg"
        image_data = render_page(page, PAGE_DIR / image_name)
        page_products = extract_products_from_skus(page, page_number, price_list)
        for product in page_products:
            product["section"] = entry["section"]
        price_groups = build_price_groups(page, page_products) if entry["showPriceOverlays"] else []
        if entry["showPriceOverlays"]:
            apply_reference_price_positions(price_groups, extract_reference_price_positions(reference_doc, page_number))
        pages.append(
            {
                "number": page_number,
                "title": find_category(page.get_text("blocks")),
                "section": entry["section"],
                "showPriceOverlays": entry["showPriceOverlays"],
                "image": image_data,
                "products": [product["id"] for product in page_products],
                "priceGroups": price_groups,
            }
        )
        products.extend(page_products)

    payload = {
        "source": str(pdf),
        "pricedReference": str(reference_pdf) if reference_pdf else "",
        "totalPagesInPdf": doc.page_count,
        "priceList": {
            "source": str(EXCEL_PATH) if EXCEL_PATH.exists() else "",
            "productCount": len(price_list),
        },
        "samplePageCount": len(pages),
        "pages": pages,
        "products": products,
    }
    json_text = json.dumps(payload, ensure_ascii=False, indent=2)
    (DATA_DIR / "catalog.json").write_text(json_text, encoding="utf-8")
    (DATA_DIR / "catalog-data.js").write_text(f"window.CATALOG_DATA = {json_text};\n", encoding="utf-8")
    excel_matches = sum(1 for product in products if product["priceSource"] == "excel")
    print(f"Generated {len(pages)} pages and {len(products)} products into {WEB_DIR}")
    print(f"Loaded {len(price_list)} Excel products; matched {excel_matches} sample catalog products")


if __name__ == "__main__":
    main()
