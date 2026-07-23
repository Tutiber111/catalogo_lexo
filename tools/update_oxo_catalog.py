from __future__ import annotations

import json
import re
from collections import defaultdict
from dataclasses import dataclass
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path

import fitz
from openpyxl import load_workbook
from PIL import Image, ImageDraw


ROOT = Path(__file__).resolve().parents[1]
WEB_DIR = ROOT / "web"
PAGE_DIR = WEB_DIR / "assets" / "pages"
DATA_JSON = WEB_DIR / "data" / "catalog.json"
DATA_JS = WEB_DIR / "data" / "catalog-data.js"
PDF_PATH = Path.home() / "Downloads" / "Catálogo OXO nuevo 2026.pdf"
PRICE_LIST_PATH = Path.home() / "Downloads" / "Lista Lexo - Julio 2026.xlsx"

BRAND = "OXO"
ASSET_VERSION = "20260723-oxo-r4"
ASSET_PREFIX = "oxo-20260722-page"
RENDER_SCALE = 1.7
OXO_FIRST_ROW = 696
OXO_LAST_ROW = 924
PRICE_RE = re.compile(r"^\$[\d.]+(?:,\d+)?$")

# The PDF still prints the retired code; the July list explicitly names its replacement.
PRINTED_SKU_ALIASES = {"11282700": "1255680"}

# This SKU appears in the PDF's POP reference chart but has no July-list row.
# It remains visible in the page image but is intentionally not orderable.
UNPRICED_PDF_SKUS = {"11234200"}


@dataclass
class PriceWord:
    text: str
    word: tuple
    index: int
    color: str
    font_size: float

    @property
    def x(self) -> float:
        return (self.word[0] + self.word[2]) / 2

    @property
    def y(self) -> float:
        return self.word[1]

    @property
    def width(self) -> float:
        return self.word[2] - self.word[0]

    @property
    def height(self) -> float:
        return self.word[3] - self.word[1]


def clean_text(value) -> str:
    return " ".join(str(value).replace("\n", " ").split()).strip()


def normalize_cell_code(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return clean_text(value)


def normalize_printed_sku(value) -> str:
    return clean_text(value).strip(".,;:()[]{}").rstrip("-").upper()


def format_price(value) -> str:
    rounded = Decimal(str(value)).quantize(Decimal("1"), rounding=ROUND_HALF_UP)
    return "$" + f"{int(rounded):,}".replace(",", ".")


def load_oxo_price_list() -> tuple[dict[str, dict], list[dict], int]:
    if not PRICE_LIST_PATH.exists():
        raise FileNotFoundError(PRICE_LIST_PATH)

    workbook = load_workbook(PRICE_LIST_PATH, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    aliases: dict[str, dict] = {}
    rows: list[dict] = []
    total_products = 0

    for row_number, row in enumerate(sheet.iter_rows(min_col=1, max_col=6, values_only=True), 1):
        code = normalize_cell_code(row[1])
        description = clean_text(row[2]) if row[2] else ""
        if code and description and isinstance(row[5], (int, float)):
            total_products += 1
        if not (OXO_FIRST_ROW <= row_number <= OXO_LAST_ROW):
            continue
        if not code or not description or not isinstance(row[5], (int, float)):
            continue

        item = {
            "sourceRow": row_number,
            "sourceCode": code,
            "description": description,
            "ean": normalize_cell_code(row[3]),
            "unitsPerCase": row[4] if isinstance(row[4], (int, float)) else None,
            "priceValue": row[5],
            "price": format_price(row[5]),
        }
        item_aliases = [part.strip() for part in re.split(r"\s*/\s*", code) if part.strip()]
        item["aliases"] = item_aliases
        rows.append(item)
        for alias in item_aliases:
            aliases[alias] = item

    workbook.close()
    return aliases, rows, total_products


def price_words_for_page(page: fitz.Page) -> list[PriceWord]:
    styles: dict[tuple[str, int], dict] = {}
    counters: dict[str, int] = defaultdict(int)
    for block in page.get_text("dict").get("blocks", []):
        for line in block.get("lines", []):
            for span in line.get("spans", []):
                text = clean_text(span.get("text", ""))
                if not PRICE_RE.fullmatch(text):
                    continue
                occurrence = counters[text]
                counters[text] += 1
                styles[(text, occurrence)] = {
                    "color": f"#{int(span.get('color', 0)):06x}",
                    "fontSize": float(span.get("size", 13)),
                }

    counters.clear()
    prices = []
    for word in page.get_text("words"):
        text = clean_text(word[4])
        if not PRICE_RE.fullmatch(text):
            continue
        occurrence = counters[text]
        counters[text] += 1
        style = styles.get((text, occurrence), {})
        prices.append(
            PriceWord(
                text=text,
                word=word,
                index=len(prices),
                color=style.get("color", "#111111"),
                font_size=style.get("fontSize", word[3] - word[1]),
            )
        )
    return prices


def title_for_page(page: fitz.Page) -> str:
    candidates = []
    for block in page.get_text("dict").get("blocks", []):
        for line in block.get("lines", []):
            spans = line.get("spans", [])
            label = clean_text(" ".join(span.get("text", "") for span in spans))
            if not label or not spans:
                continue
            y0 = min(span["bbox"][1] for span in spans)
            if y0 > page.rect.height * 0.30 or PRICE_RE.search(label):
                continue
            if not any(ch.isalpha() for ch in label) or label.lower() in {"oxo", "thoughtfully yours."}:
                continue
            font_size = max(float(span.get("size", 0)) for span in spans)
            width = max(span["bbox"][2] for span in spans) - min(span["bbox"][0] for span in spans)
            candidates.append((font_size, -y0, width, label))
    return max(candidates)[3] if candidates else "Catálogo"


def normalized_hotspot(page: fitz.Page, word: tuple) -> dict[str, float]:
    pad_x = 3
    pad_y = 2
    x = max(0, word[0] - pad_x) / page.rect.width
    y = max(0, word[1] - pad_y) / page.rect.height
    return {
        "x": round(x, 7),
        "y": round(y, 7),
        "w": round((word[2] - word[0] + pad_x * 2) / page.rect.width, 7),
        "h": round((word[3] - word[1] + pad_y * 2) / page.rect.height, 7),
    }


def nearest_price(sku_word: tuple, prices: list[PriceWord]) -> PriceWord | None:
    if not prices:
        return None
    sx = (sku_word[0] + sku_word[2]) / 2
    sy = sku_word[1]
    below = [price for price in prices if -3 <= price.y - sy <= 115 and abs(price.x - sx) <= 180]
    if below:
        return min(below, key=lambda price: abs(price.x - sx) + max(0, price.y - sy) * 0.35)
    if len(prices) == 1:
        return prices[0]
    nearby = [price for price in prices if abs(price.y - sy) <= 145 and abs(price.x - sx) <= 180]
    if nearby:
        return min(nearby, key=lambda price: abs(price.x - sx) + abs(price.y - sy) * 0.45)
    return None


def price_position(page: fitz.Page, price: PriceWord | None, sku_word: tuple) -> dict[str, float]:
    if price:
        return {"x": round(price.x / page.rect.width, 7), "y": round(price.y / page.rect.height, 7)}
    return {
        "x": round(((sku_word[0] + sku_word[2]) / 2) / page.rect.width, 7),
        "y": round(min(0.95, (sku_word[3] + 5) / page.rect.height), 7),
    }


def price_style(page: fitz.Page, price: PriceWord) -> dict:
    return {
        "fontSize": round((price.font_size / page.rect.width) * 100, 3),
        "fontSizeUnit": "cqw",
        "fontWeight": 800,
        "minWidth": 0,
        "minHeight": 0,
        "padX": 0,
        "padY": 0,
        "radius": 1,
        "shadow": "none",
        "color": "#111111",
        "background": "#f5f5f5",
        "borderColor": "rgba(215, 25, 32, 0.10)",
    }


def price_cover(page: fitz.Page, price: PriceWord, display_price: str) -> dict[str, float]:
    estimated_text_width = price.font_size * 0.58 * len(display_price)
    width = max(price.width + 2, estimated_text_width + 3)
    return {
        "w": round(width / page.rect.width, 7),
        "h": round((price.height + 2) / page.rect.height, 7),
    }


def erase_printed_prices(image: Image.Image, prices: list[PriceWord]) -> None:
    draw = ImageDraw.Draw(image)
    for price in prices:
        x0, y0, x1, y1 = price.word[:4]
        draw.rectangle(
            (
                int((x0 - 2) * RENDER_SCALE),
                int((y0 - 1) * RENDER_SCALE),
                int((x1 + 2) * RENDER_SCALE),
                int((y1 + 2) * RENDER_SCALE),
            ),
            fill=(245, 245, 245),
        )


def render_page(page: fitz.Page, destination: Path, prices: list[PriceWord]) -> dict:
    pix = page.get_pixmap(matrix=fitz.Matrix(RENDER_SCALE, RENDER_SCALE), alpha=False)
    image = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
    erase_printed_prices(image, prices)
    destination.parent.mkdir(parents=True, exist_ok=True)
    image.save(destination, "JPEG", quality=89, optimize=True)
    return {
        "src": f"{destination.relative_to(WEB_DIR).as_posix()}?v={ASSET_VERSION}",
        "width": pix.width,
        "height": pix.height,
    }


def product_hits(page: fitz.Page, price_aliases: dict[str, dict]) -> tuple[list[dict], list[dict]]:
    hits = []
    unpriced = []
    seen = set()
    for word in page.get_text("words"):
        printed_sku = normalize_printed_sku(word[4])
        if printed_sku in UNPRICED_PDF_SKUS:
            unpriced.append({"sku": printed_sku, "word": word})
            continue
        lookup_sku = PRINTED_SKU_ALIASES.get(printed_sku, printed_sku)
        price_data = price_aliases.get(lookup_sku)
        if not price_data or (word[3] - word[1]) < 11.5:
            continue
        key = (lookup_sku, round(word[0], 2), round(word[1], 2))
        if key in seen:
            continue
        seen.add(key)
        hits.append(
            {
                "sku": lookup_sku if printed_sku in PRINTED_SKU_ALIASES else printed_sku,
                "printedSku": printed_sku,
                "word": word,
                "priceData": price_data,
            }
        )
    return hits, unpriced


def build_oxo_pages(start_page: int, price_aliases: dict[str, dict]) -> tuple[list[dict], list[dict], set[int], set[str]]:
    if not PDF_PATH.exists():
        raise FileNotFoundError(PDF_PATH)

    document = fitz.open(PDF_PATH)
    pages = []
    products = []
    matched_rows: set[int] = set()
    printed_skus: set[str] = set()

    for page_index, page in enumerate(document):
        source_page = page_index + 1
        app_page = start_page + page_index
        title = title_for_page(page)
        prices = price_words_for_page(page)
        hits, unpriced_hits = product_hits(page, price_aliases)

        # PDF-only products have no authoritative spreadsheet price, so no ordering
        # hotspot or price overlay is created for them.
        unpriced_prices = {nearest_price(hit["word"], prices).index for hit in unpriced_hits if nearest_price(hit["word"], prices)}
        image = render_page(page, PAGE_DIR / f"{ASSET_PREFIX}-{source_page:03d}.jpg", prices)
        page_products = []
        price_members: dict[int, list[dict]] = defaultdict(list)

        for hit in hits:
            printed_skus.add(hit["printedSku"])
            matched_rows.add(hit["priceData"]["sourceRow"])
            nearest = nearest_price(hit["word"], prices)
            product_id = f"oxo-p{source_page:03d}-{len(page_products) + 1}"
            product = {
                "id": product_id,
                "page": app_page,
                "sku": hit["sku"],
                "skus": [hit["sku"]],
                "name": hit["priceData"]["description"],
                "category": title,
                "price": hit["priceData"]["price"],
                "pdfPrice": nearest.text if nearest else "",
                "priceSource": "excel-july-2026",
                "ean": hit["priceData"]["ean"],
                "unitsPerCase": hit["priceData"]["unitsPerCase"],
                "sizeLabel": "",
                "hotspot": normalized_hotspot(page, hit["word"]),
                "hotspotStyle": {"borderColor": "rgba(215, 25, 32, 0.42)"},
                "pricePosition": price_position(page, nearest, hit["word"]),
                "section": BRAND,
                "sourcePage": source_page,
                "hotspotSource": "oxo-20260722-sku-text",
            }
            if hit["printedSku"] != hit["sku"]:
                product["printedSku"] = hit["printedSku"]
            page_products.append(product)
            if nearest and nearest.index not in unpriced_prices:
                price_members[nearest.index].append(product)

        price_groups = []
        for price_index, members in sorted(price_members.items()):
            printed_price = prices[price_index]
            distinct_prices = list(dict.fromkeys(product["price"] for product in members))
            display_price = " / ".join(distinct_prices)
            price_groups.append(
                {
                    "id": f"oxo-pg{source_page:03d}-{len(price_groups) + 1}",
                    "page": app_page,
                    "label": title if len(members) == 1 else f"{title} - {len(members)} productos",
                    "price": display_price,
                    "productIds": [product["id"] for product in members],
                    "position": price_position(page, printed_price, members[0]["hotspot"]),
                    "positionSource": "oxo-20260722-pdf",
                    "cover": price_cover(page, printed_price, display_price),
                    "variant": "pdf-regular",
                    "style": price_style(page, printed_price),
                    "pdfPriceHeight": round(printed_price.height, 3),
                    "pdfPriceColor": printed_price.color,
                }
            )

        pages.append(
            {
                "number": app_page,
                "title": title,
                "section": BRAND,
                "showPriceOverlays": bool(price_groups),
                "image": image,
                "products": [product["id"] for product in page_products],
                "priceGroups": price_groups,
                "sourcePage": source_page,
            }
        )
        products.extend(page_products)

    document.close()
    return pages, products, matched_rows, printed_skus


def write_catalog(catalog: dict) -> None:
    json_text = json.dumps(catalog, ensure_ascii=False, indent=2) + "\n"
    DATA_JSON.write_text(json_text, encoding="utf-8")
    DATA_JS.write_text(f"window.CATALOG_DATA = {json_text.rstrip()};\n", encoding="utf-8")


def update_cache_versions() -> None:
    replacements = {
        WEB_DIR / "index.html": [
            (r"((?:styles|app)\.css\?v=)[^\"<]+", rf"\g<1>{ASSET_VERSION}"),
            (r"((?:app|supabase-client)\.js\?v=)[^\"<]+", rf"\g<1>{ASSET_VERSION}"),
            (r"(data/catalog-data\.js\?v=)[^\"<]+", rf"\g<1>{ASSET_VERSION}"),
        ],
        WEB_DIR / "service-worker.js": [
            (r'lexo-catalog-v[^\"]+', f"lexo-catalog-v{ASSET_VERSION}"),
            (r"(\./(?:styles|app)\.css\?v=)[^\"']+", rf"\g<1>{ASSET_VERSION}"),
            (r"(\./app\.js\?v=)[^\"']+", rf"\g<1>{ASSET_VERSION}"),
            (r"(\./supabase-client\.js\?v=)[^\"']+", rf"\g<1>{ASSET_VERSION}"),
            (r"(\./data/catalog-data\.js\?v=)[^\"']+", rf"\g<1>{ASSET_VERSION}"),
        ],
    }
    for path, rules in replacements.items():
        text = path.read_text(encoding="utf-8")
        for pattern, replacement in rules:
            text = re.sub(pattern, replacement, text)
        path.write_text(text, encoding="utf-8")


def validate_catalog(catalog: dict) -> None:
    page_numbers = [page["number"] for page in catalog["pages"]]
    if page_numbers != list(range(1, len(page_numbers) + 1)):
        raise ValueError("Catalog pages are not contiguous")

    product_ids = [product["id"] for product in catalog["products"]]
    if len(product_ids) != len(set(product_ids)):
        raise ValueError("Duplicate product IDs detected")
    known_ids = set(product_ids)
    missing_refs = [
        product_id
        for page in catalog["pages"]
        for product_id in page.get("products", [])
        if product_id not in known_ids
    ]
    if missing_refs:
        raise ValueError(f"Missing product references: {missing_refs[:20]}")

    grouped = [
        product_id
        for page in catalog["pages"]
        for group in page.get("priceGroups", [])
        for product_id in group.get("productIds", [])
    ]
    if any(product_id not in known_ids for product_id in grouped):
        raise ValueError("A price group references an unknown product")
    if any(product.get("skus") != [product.get("sku")] for product in catalog["products"]):
        raise ValueError("Related SKU groups must remain disabled")


def main() -> None:
    catalog = json.loads(DATA_JSON.read_text(encoding="utf-8"))
    old_pages = [page for page in catalog["pages"] if page.get("section") == BRAND]
    old_products = [product for product in catalog["products"] if product.get("section") == BRAND]
    if not old_pages:
        raise ValueError("Could not locate the existing OXO section")

    old_start = min(page["number"] for page in old_pages)
    old_end = max(page["number"] for page in old_pages)
    old_skus = {product["sku"] for product in old_products}
    price_aliases, price_rows, total_price_products = load_oxo_price_list()
    new_pages, new_products, matched_rows, printed_skus = build_oxo_pages(old_start, price_aliases)
    delta = len(new_pages) - len(old_pages)

    before_pages = [page for page in catalog["pages"] if page.get("section") != BRAND and page["number"] < old_start]
    after_pages = [page for page in catalog["pages"] if page.get("section") != BRAND and page["number"] > old_end]
    for page in after_pages:
        page["number"] += delta
        for group in page.get("priceGroups", []):
            if "page" in group:
                group["page"] += delta

    before_products = [
        product for product in catalog["products"] if product.get("section") != BRAND and product["page"] < old_start
    ]
    after_products = [
        product for product in catalog["products"] if product.get("section") != BRAND and product["page"] > old_end
    ]
    for product in after_products:
        product["page"] += delta

    catalog["pages"] = before_pages + new_pages + after_pages
    catalog["products"] = before_products + new_products + after_products
    catalog.setdefault("sources", {})[BRAND] = str(PDF_PATH)
    catalog.setdefault("sourcePageCounts", {})[BRAND] = len(new_pages)
    catalog["priceList"] = {"source": str(PRICE_LIST_PATH), "productCount": total_price_products}
    catalog["totalPagesInPdf"] = len(catalog["pages"])
    catalog["samplePageCount"] = len(catalog["pages"])
    catalog["assetVersion"] = ASSET_VERSION

    validate_catalog(catalog)
    write_catalog(catalog)
    update_cache_versions()

    new_skus = {product["sku"] for product in new_products}
    missing_rows = [row for row in price_rows if row["sourceRow"] not in matched_rows]
    report = {
        "oldPages": len(old_pages),
        "newPages": len(new_pages),
        "oldPlacements": len(old_products),
        "newPlacements": len(new_products),
        "oldUniqueSkus": len(old_skus),
        "newUniqueSkus": len(new_skus),
        "newSkusVsOld": sorted(new_skus - old_skus),
        "oldSkusNotOrderableInReplacement": sorted(old_skus - new_skus),
        "priceListRowsMissingFromPdf": [
            {"sku": row["sourceCode"], "name": row["description"], "row": row["sourceRow"]}
            for row in missing_rows
        ],
        "pdfSkusMissingFromPriceList": sorted(UNPRICED_PDF_SKUS),
        "printedSkuAliases": PRINTED_SKU_ALIASES,
        "unpricedPdfSkus": sorted(UNPRICED_PDF_SKUS),
        "printedMatchedSkus": sorted(printed_skus),
        "catalogPages": len(catalog["pages"]),
        "catalogProducts": len(catalog["products"]),
    }
    report_path = ROOT / "tmp" / "oxo-update-report.json"
    report_path.parent.mkdir(parents=True, exist_ok=True)
    report_path.write_text(json.dumps(report, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(json.dumps(report, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
